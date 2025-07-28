"""
Модуль экспорта результатов в Excel
"""

import logging
from typing import Dict, List, Optional, Any, Union
import json
from dataclasses import dataclass
import pandas as pd
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils.dataframe import dataframe_to_rows
from openpyxl.worksheet.worksheet import Worksheet
from pathlib import Path
from datetime import datetime

logger = logging.getLogger(__name__)


@dataclass
class ExcelExportConfig:
    """Конфигурация экспорта в Excel"""
    include_statistics: bool = True
    include_metadata: bool = True
    auto_adjust_columns: bool = True
    add_filters: bool = True
    highlight_high_similarity: bool = True
    similarity_threshold: float = 0.8
    max_results_per_query: int = 50
    include_processing_details: bool = True


# Alias for backward compatibility
ExportConfig = ExcelExportConfig


class ExcelExporter:
    """Класс для экспорта результатов в Excel"""

    def __init__(self, config: ExcelExportConfig = None):
        self.config = config or ExcelExportConfig()
        logger.info("ExcelExporter initialized")
    
    def export_search_results(self,
                            results: Dict[str, List[Dict[str, Any]]],
                            output_path: str,
                            metadata: Dict[str, Any] = None) -> str:
        """
        Экспорт результатов поиска в Excel

        Args:
            results: Словарь с результатами поиска {query: [results]}
            output_path: Путь к выходному файлу
            metadata: Дополнительные метаданные

        Returns:
            Путь к созданному файлу
        """
        try:
            # Создаем новую книгу Excel
            wb = Workbook()
            
            # Удаляем стандартный лист
            wb.remove(wb.active)
            
            # Создаем основной лист с результатами
            self._create_results_sheet(wb, results)
            
            # Создаем лист со статистикой
            if self.config.include_statistics:
                self._create_statistics_sheet(wb, results)
            
            # Создаем лист с метаданными
            if metadata:
                self._create_metadata_sheet(wb, metadata)
            
            # Сохраняем файл
            wb.save(output_path)
            logger.info(f"Results exported to {output_path}")

            return output_path
            
        except Exception as e:
            logger.error(f"Error exporting to Excel: {e}")
            raise
    
    def _create_results_sheet(self, wb: Workbook, results: Dict[str, List[Dict[str, Any]]]):
        """Создание листа с результатами поиска"""
        ws = wb.create_sheet("Результаты поиска")
        
        # Подготовка данных
        all_results = []
        for query, query_results in results.items():
            for i, result in enumerate(query_results[:self.config.max_results_per_query]):
                row_data = {
                    'Запрос': query,
                    'Ранг': i + 1,
                    'Найденное наименование': result.get('document', ''),
                    'ID документа': result.get('document_id', ''),
                    'Оценка схожести': result.get('similarity_score', result.get('combined_score', 0)),
                    'Тип поиска': self._determine_search_type(result),
                    'Дополнительные метрики': self._format_additional_metrics(result)
                }
                
                # Добавляем извлеченные параметры если есть
                if 'extracted_parameters' in result:
                    row_data['Извлеченные параметры'] = self._format_parameters(result['extracted_parameters'])
                    # Добавляем дополнительную колонку с исходным текстом параметров
                    row_data['Исходный текст параметров'] = self._format_parameters_source_text(result['extracted_parameters'])
                    # Добавляем структурированные параметры в JSON формате
                    row_data['Структурированные параметры'] = self._format_parameters_json(result['extracted_parameters'])
                
                all_results.append(row_data)
        
        # Создаем DataFrame
        df = pd.DataFrame(all_results)
        
        if df.empty:
            ws.append(['Нет результатов для экспорта'])
            return
        
        # Записываем данные в лист
        for r in dataframe_to_rows(df, index=False, header=True):
            ws.append(r)
        
        # Применяем стили
        self._apply_results_styling(ws, df)
        
        # Настраиваем колонки
        if self.config.auto_adjust_columns:
            self._adjust_column_widths(ws)
        
        # Добавляем фильтры
        if self.config.add_filters:
            ws.auto_filter.ref = ws.dimensions
    
    def _create_statistics_sheet(self, wb: Workbook, results: Dict[str, List[Dict[str, Any]]]):
        """Создание листа со статистикой"""
        ws = wb.create_sheet("Статистика")
        
        # Общая статистика
        total_queries = len(results)
        total_results = sum(len(query_results) for query_results in results.values())
        avg_results_per_query = total_results / total_queries if total_queries > 0 else 0
        
        # Статистика по качеству
        high_quality_results = 0
        similarity_scores = []
        
        for query_results in results.values():
            for result in query_results:
                score = result.get('similarity_score', result.get('combined_score', 0))
                similarity_scores.append(score)
                if score >= self.config.similarity_threshold:
                    high_quality_results += 1
        
        avg_similarity = sum(similarity_scores) / len(similarity_scores) if similarity_scores else 0
        
        # Записываем статистику
        stats_data = [
            ['Метрика', 'Значение'],
            ['Общее количество запросов', total_queries],
            ['Общее количество результатов', total_results],
            ['Среднее количество результатов на запрос', round(avg_results_per_query, 2)],
            ['Результаты высокого качества', high_quality_results],
            ['Процент высококачественных результатов', f"{(high_quality_results/total_results*100):.1f}%" if total_results > 0 else "0%"],
            ['Средняя оценка схожести', round(avg_similarity, 3)],
            ['Минимальная оценка схожести', round(min(similarity_scores), 3) if similarity_scores else 0],
            ['Максимальная оценка схожести', round(max(similarity_scores), 3) if similarity_scores else 0],
            ['Дата экспорта', datetime.now().strftime('%Y-%m-%d %H:%M:%S')]
        ]
        
        for row in stats_data:
            ws.append(row)
        
        # Применяем стили к статистике
        self._apply_statistics_styling(ws)
    
    def _create_metadata_sheet(self, wb: Workbook, metadata: Dict[str, Any]):
        """Создание листа с метаданными"""
        ws = wb.create_sheet("Метаданные")
        
        # Записываем метаданные
        ws.append(['Параметр', 'Значение'])
        
        for key, value in metadata.items():
            if isinstance(value, (dict, list)):
                value = str(value)
            ws.append([key, value])
        
        # Применяем стили
        self._apply_metadata_styling(ws)
    
    def _determine_search_type(self, result: Dict[str, Any]) -> str:
        """Определение типа поиска по результату"""
        # Проверяем явно указанный метод поиска (приоритет)
        if 'search_method' in result:
            method = result['search_method'].lower()
            if method == 'hybrid':
                return 'Гибридный'
            elif method == 'semantic':
                return 'Семантический'
            elif method == 'fuzzy':
                return 'Нечеткий'

        # Fallback: определяем по наличию скоров
        # Гибридный поиск имеет и similarity_score и combined_score
        if ('similarity_score' in result and 'combined_score' in result) or 'hybrid_score' in result:
            return 'Гибридный'
        elif 'similarity_score' in result:
            return 'Семантический'
        elif 'combined_score' in result or 'fuzzy_score' in result:
            return 'Нечеткий'
        else:
            return 'Неизвестный'
    
    def _format_additional_metrics(self, result: Dict[str, Any]) -> str:
        """Форматирование дополнительных метрик"""
        metrics = []

        # Гибридные метрики
        if 'hybrid_score' in result:
            metrics.append(f"Гибридный: {result['hybrid_score']:.3f}")

        # Семантические метрики
        if 'similarity_score' in result:
            metrics.append(f"Семантический: {result['similarity_score']:.3f}")

        if 'cosine_score' in result:
            metrics.append(f"Косинус: {result['cosine_score']:.3f}")

        # Нечеткие метрики
        if 'combined_score' in result:
            metrics.append(f"Комбинированный: {result['combined_score']:.3f}")

        if 'fuzzy_score' in result:
            # Проверяем тип значения (может быть int или float)
            fuzzy_val = result['fuzzy_score']
            if isinstance(fuzzy_val, (int, float)):
                metrics.append(f"Нечеткий: {fuzzy_val:.3f}")
            else:
                metrics.append(f"Нечеткий: {fuzzy_val}")

        if 'levenshtein_score' in result:
            metrics.append(f"Левенштейн: {result['levenshtein_score']:.3f}")

        # Дополнительные метрики для гибридного поиска
        if 'semantic_score' in result:
            metrics.append(f"Сем.компонент: {result['semantic_score']:.3f}")

        if 'fuzzy_rank' in result and result['fuzzy_rank'] is not None:
            metrics.append(f"Ранг нечеткий: {result['fuzzy_rank']}")

        if 'semantic_rank' in result and result['semantic_rank'] is not None:
            metrics.append(f"Ранг семантический: {result['semantic_rank']}")

        # Информация о стратегии комбинирования
        if 'combination_strategy' in result:
            strategy_names = {
                'weighted_sum': 'Взвешенная сумма',
                'rank_fusion': 'Слияние рангов',
                'cascade': 'Каскадный'
            }
            strategy = strategy_names.get(result['combination_strategy'], result['combination_strategy'])
            metrics.append(f"Стратегия: {strategy}")

        return '; '.join(metrics)
    
    def _format_parameters(self, parameters) -> str:
        """Форматирование извлеченных параметров"""
        if not parameters:
            return ''

        # Handle case where parameters is a JSON string (serialized format)
        if isinstance(parameters, str):
            try:
                parameters = json.loads(parameters)
            except (json.JSONDecodeError, TypeError):
                # If it's not valid JSON, return the string as-is
                return str(parameters)

        # Handle case where parameters is not a list
        if not isinstance(parameters, list):
            return str(parameters)

        formatted = []
        for param in parameters:
            # Handle case where param is not a dictionary
            if not isinstance(param, dict):
                formatted.append(str(param))
                continue

            name = param.get('name', '')
            value = param.get('value', '')
            unit = param.get('unit', '')

            param_str = f"{name}: {value}"
            if unit:
                param_str += f" {unit}"

            formatted.append(param_str)

        return '; '.join(formatted)

    def _format_parameters_source_text(self, parameters) -> str:
        """Форматирование исходного текста параметров"""
        if not parameters:
            return ''

        # Handle case where parameters is a JSON string (serialized format)
        if isinstance(parameters, str):
            try:
                parameters = json.loads(parameters)
            except (json.JSONDecodeError, TypeError):
                return str(parameters)

        # Handle case where parameters is not a list
        if not isinstance(parameters, list):
            return str(parameters)

        source_texts = []
        for param in parameters:
            if not isinstance(param, dict):
                continue

            source_text = param.get('source_text', '')
            if source_text:
                source_texts.append(source_text)

        return '; '.join(source_texts)

    def _format_parameters_json(self, parameters) -> str:
        """Форматирование параметров в JSON формате"""
        if not parameters:
            return ''

        # Handle case where parameters is a JSON string (serialized format)
        if isinstance(parameters, str):
            try:
                parameters = json.loads(parameters)
            except (json.JSONDecodeError, TypeError):
                return str(parameters)

        # Handle case where parameters is not a list
        if not isinstance(parameters, list):
            return str(parameters)

        # Create structured parameter data
        structured_params = []
        for param in parameters:
            if not isinstance(param, dict):
                continue

            structured_param = {
                'name': param.get('name', ''),
                'value': param.get('value', ''),
                'unit': param.get('unit', ''),
                'confidence': param.get('confidence', 0.0),
                'parameter_type': param.get('parameter_type', ''),
                'position': param.get('position', [])
            }
            structured_params.append(structured_param)

        # Return as formatted JSON string
        try:
            return json.dumps(structured_params, ensure_ascii=False, indent=None)
        except Exception:
            return str(structured_params)

    def _apply_results_styling(self, ws: Worksheet, df: pd.DataFrame):
        """Применение стилей к листу результатов"""
        # Стиль заголовка
        header_font = Font(bold=True, color="FFFFFF")
        header_fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
        header_alignment = Alignment(horizontal="center", vertical="center")
        
        # Применяем стили к заголовку
        for cell in ws[1]:
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = header_alignment
        
        # Подсветка высококачественных результатов
        if self.config.highlight_high_similarity:
            similarity_col = None
            for i, col_name in enumerate(df.columns, 1):
                if 'схожести' in col_name.lower():
                    similarity_col = i
                    break
            
            if similarity_col:
                high_quality_fill = PatternFill(start_color="C6EFCE", end_color="C6EFCE", fill_type="solid")
                
                for row in range(2, ws.max_row + 1):
                    cell = ws.cell(row=row, column=similarity_col)
                    if isinstance(cell.value, (int, float)) and cell.value >= self.config.similarity_threshold:
                        for col in range(1, ws.max_column + 1):
                            ws.cell(row=row, column=col).fill = high_quality_fill
    
    def _apply_statistics_styling(self, ws: Worksheet):
        """Применение стилей к листу статистики"""
        # Стиль заголовка
        header_font = Font(bold=True, color="FFFFFF")
        header_fill = PatternFill(start_color="70AD47", end_color="70AD47", fill_type="solid")
        
        # Применяем к первой строке
        for cell in ws[1]:
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = Alignment(horizontal="center")
        
        # Стиль для значений
        for row in range(2, ws.max_row + 1):
            ws.cell(row=row, column=1).font = Font(bold=True)
    
    def _apply_metadata_styling(self, ws: Worksheet):
        """Применение стилей к листу метаданных"""
        # Стиль заголовка
        header_font = Font(bold=True, color="FFFFFF")
        header_fill = PatternFill(start_color="FFC000", end_color="FFC000", fill_type="solid")
        
        for cell in ws[1]:
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = Alignment(horizontal="center")
    
    def _adjust_column_widths(self, ws: Worksheet):
        """Автоматическая настройка ширины колонок"""
        for column in ws.columns:
            max_length = 0
            column_letter = column[0].column_letter
            
            for cell in column:
                try:
                    if len(str(cell.value)) > max_length:
                        max_length = len(str(cell.value))
                except:
                    pass
            
            adjusted_width = min(max_length + 2, 50)  # Максимум 50 символов
            ws.column_dimensions[column_letter].width = adjusted_width
    
    def export_comparison_table(self, 
                              original_items: List[Dict[str, Any]], 
                              analog_results: Dict[str, List[Dict[str, Any]]], 
                              filepath: str) -> str:
        """
        Экспорт таблицы сравнения оригинальных позиций с найденными аналогами
        
        Args:
            original_items: Список оригинальных позиций
            analog_results: Результаты поиска аналогов
            filepath: Путь к выходному файлу
            
        Returns:
            Путь к созданному файлу
        """
        try:
            wb = Workbook()
            ws = wb.active
            ws.title = "Сравнение с аналогами"
            
            # Заголовки
            headers = [
                'ID оригинала', 'Оригинальное наименование', 
                'Лучший аналог', 'Оценка схожести', 'Тип связи',
                'Количество найденных аналогов', 'Все аналоги'
            ]
            ws.append(headers)
            
            # Заполняем данные
            for item in original_items:
                item_id = item.get('id', '')
                original_name = item.get('name', '')
                
                # Получаем результаты для этой позиции
                results = analog_results.get(original_name, [])
                
                if results:
                    best_analog = results[0]
                    best_name = best_analog.get('document', '')
                    best_score = best_analog.get('similarity_score', best_analog.get('combined_score', 0))
                    relation_type = self._determine_relation_type(best_score)
                    
                    all_analogs = '; '.join([r.get('document', '')[:50] + '...' if len(r.get('document', '')) > 50 
                                           else r.get('document', '') for r in results[:5]])
                else:
                    best_name = 'Не найдено'
                    best_score = 0
                    relation_type = 'Нет аналогов'
                    all_analogs = ''
                
                row = [
                    item_id, original_name, best_name, 
                    round(best_score, 3), relation_type, 
                    len(results), all_analogs
                ]
                ws.append(row)
            
            # Применяем стили
            self._apply_comparison_styling(ws)
            
            # Настраиваем колонки
            if self.config.auto_adjust_columns:
                self._adjust_column_widths(ws)
            
            # Добавляем фильтры
            if self.config.add_filters:
                ws.auto_filter.ref = ws.dimensions
            
            wb.save(filepath)
            logger.info(f"Comparison table exported to {filepath}")
            
            return filepath
            
        except Exception as e:
            logger.error(f"Error exporting comparison table: {e}")
            raise
    
    def _determine_relation_type(self, score: float) -> str:
        """Определение типа связи по оценке схожести"""
        if score >= 0.9:
            return 'Точный аналог'
        elif score >= 0.7:
            return 'Близкий аналог'
        elif score >= 0.5:
            return 'Возможный аналог'
        else:
            return 'Слабое соответствие'
    
    def _apply_comparison_styling(self, ws: Worksheet):
        """Применение стилей к таблице сравнения"""
        # Стиль заголовка
        header_font = Font(bold=True, color="FFFFFF")
        header_fill = PatternFill(start_color="5B9BD5", end_color="5B9BD5", fill_type="solid")
        
        for cell in ws[1]:
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = Alignment(horizontal="center", vertical="center")
        
        # Условное форматирование для оценок схожести
        score_col = 4  # Колонка с оценкой схожести
        
        for row in range(2, ws.max_row + 1):
            score_cell = ws.cell(row=row, column=score_col)
            
            if isinstance(score_cell.value, (int, float)):
                if score_cell.value >= 0.8:
                    fill_color = "C6EFCE"  # Зеленый
                elif score_cell.value >= 0.6:
                    fill_color = "FFEB9C"  # Желтый
                else:
                    fill_color = "FFC7CE"  # Красный
                
                score_cell.fill = PatternFill(start_color=fill_color, end_color=fill_color, fill_type="solid")
