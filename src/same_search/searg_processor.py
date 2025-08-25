import pandas as pd
from dataclasses import dataclass
from pathlib import Path
from typing import Dict, List, Any
from datetime import datetime
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment
from openpyxl.utils.dataframe import dataframe_to_rows
import asyncio

from src.same_api.export import ExcelExporter, ExportConfig
from src.same_search.categorization import CategoryClassifier, CategoryClassifierConfig
from src.same_clear.search import SearchResult, SearchConfig, TokenSearchEngine
from src.same_clear import PreprocessorConfig
from src.same_search.search_interface import SAMeSearchInterface
from src.same.analog_search_engine import AnalogSearchEngine
from src.same_search.hybrid_dbscan_search import HybridDBSCANSearchEngine, HybridDBSCANConfig
from src.same_search.optimized_dbscan_search import OptimizedHybridDBSCANSearchEngine, OptimizedDBSCANConfig

import logging

logger = logging.getLogger(__name__)


@dataclass
class AnalogSearchConfig:
    """Конфигурация системы поиска аналогов"""
    # Конфигурации компонентов
    preprocessor_config: PreprocessorConfig = None
    export_config: ExportConfig = None
    
    # Параметры поиска
    search_method: str = "hybrid"  
    similarity_threshold: float = 0.6
    max_results_per_query: int = 100
    
    # Параметры обработки
    batch_size: int = 100
    enable_parameter_extraction: bool = True
    
    # Пути к данным
    data_dir: Path = Path("data")
    models_dir: Path = Path("models")
    output_dir: Path = Path("data/output")

class AnalogSearchProcessor:
    """Основной класс для обработки каталога и поиска аналогов"""

    def __init__(self, search_method: str = "hybrid", similarity_threshold: float = 0.6, 
                 use_extended_search: bool = True, max_excel_results: int = 1000000):
        """
        Инициализация процессора

        Args:
            search_method: Метод поиска (fuzzy, semantic, hybrid, extended_hybrid, token_id,
                          prefix, inverted_index, tfidf, lsh, spatial, advanced_hybrid)
            similarity_threshold: Порог схожести для фильтрации результатов
            use_extended_search: Использовать новую систему расширенного поиска
            max_excel_results: Максимальное количество результатов для записи в Excel
        """
        self.search_method = search_method
        self.similarity_threshold = similarity_threshold
        self.use_extended_search = use_extended_search
        self.max_excel_results = max_excel_results

        # Формат входного файла (определяется при валидации)
        self.input_format = 'unknown'

        # Новая система поиска
        self.extended_search_interface = None

        # Новая система поиска по токенам
        self.token_search_engine = None
        self.use_token_search = self._is_token_search_method(search_method)

        # Hybrid DBSCAN поисковый движок
        self.hybrid_dbscan_engine = None
        self.use_hybrid_dbscan = search_method == "hybrid_dbscan"

        # Optimized DBSCAN поисковый движок
        self.optimized_dbscan_engine = None
        self.use_optimized_dbscan = search_method == "optimized_dbscan"
        
        # Конфигурация предобработчика (отключаем параллельную обработку для больших датасетов)
        self.preprocessor_config = PreprocessorConfig(
            save_intermediate_steps=False,  # Отключаем для экономии памяти
            batch_size=500,  # Уменьшаем размер батча
            enable_parallel_processing=False,  # Отключаем параллельную обработку
            max_workers=1,
            parallel_threshold=10000,  # Увеличиваем порог
            chunk_size=25  # Уменьшаем размер чанка
        )
        
        # Конфигурация поискового движка
        self.search_config = AnalogSearchConfig(
            preprocessor_config=self.preprocessor_config,
            search_method=search_method,
            similarity_threshold=similarity_threshold,
            max_results_per_query=50,
            batch_size=100,
            enable_parameter_extraction=True
        )

        # Конфигурация поиска по токенам
        self.token_search_config = SearchConfig(
            # Базовые настройки
            token_id_weight=0.6,
            semantic_weight=0.4,
            similarity_threshold=similarity_threshold,
            max_results=50,

            # Включаем все методы поиска
            enable_trie_search=True,
            trie_weight=0.3,
            trie_min_prefix_length=2,

            enable_inverted_index=True,
            inverted_index_weight=0.4,

            enable_tfidf_search=True,
            tfidf_weight=0.35,
            tfidf_max_features=10000,
            tfidf_ngram_range=(1, 3),

            enable_lsh_search=True,
            lsh_weight=0.25,
            lsh_threshold=0.6,
            lsh_num_perm=128,

            enable_spatial_search=True,
            spatial_weight=0.3,
            faiss_index_type="flat",

            # Продвинутые методы (отключены по умолчанию для производительности)
            enable_advanced_embeddings=False,
            enable_graph_search=False,

            boost_technical_terms=True,
            enable_fuzzy_matching=True
        )
        
        # Инициализация классификатора категорий
        self.category_classifier = CategoryClassifier(
            CategoryClassifierConfig(
                use_keyword_matching=True,
                use_pattern_matching=True,
                min_confidence=0.3
            )
        )
        
        self.search_engine = None

        # Инициализация новой системы поиска если включена
        if self.use_extended_search:
            logger.info("Initializing extended search system...")
            try:
                self.extended_search_interface = SAMeSearchInterface()
                logger.info("Extended search interface created")
            except Exception as e:
                logger.warning(f"Failed to initialize extended search: {e}")
                self.use_extended_search = False
                self.extended_search_interface = None

        # Инициализация системы поиска по токенам если нужна
        if self.use_token_search:
            logger.info("Initializing token search system...")
            try:
                self.token_search_engine = TokenSearchEngine(self.token_search_config)
                logger.info("Token search engine created")
            except Exception as e:
                logger.warning(f"Failed to initialize token search: {e}")
                self.use_token_search = False
                self.token_search_engine = None

        logger.info(f"AnalogSearchProcessor initialized (extended_search: {self.use_extended_search}, token_search: {self.use_token_search})")

    def _is_token_search_method(self, method: str) -> bool:
        """
        Определяет, является ли метод поиска методом поиска по токенам

        Args:
            method: Название метода поиска

        Returns:
            True если это метод поиска по токенам
        """
        token_search_methods = {
            'token_id', 'prefix', 'inverted_index', 'tfidf',
            'lsh', 'spatial', 'advanced_hybrid'
        }
        return method in token_search_methods

    def _is_hybrid_dbscan_method(self, method: str) -> bool:
        """Проверяет, является ли метод hybrid DBSCAN поиском"""
        return method == "hybrid_dbscan"

    def search_by_tokens(self, query: str, method: str = None, top_k: int = 10) -> List[SearchResult]:
        """
        Поиск аналогов с использованием системы поиска по токенам

        Args:
            query: Поисковый запрос
            method: Метод поиска (если None, используется self.search_method)
            top_k: Количество результатов

        Returns:
            Список результатов поиска
        """
        if not self.token_search_engine:
            logger.error("Token search engine not initialized")
            return []

        search_method = method or self.search_method

        # Добавляем диагностику для spatial поиска
        if search_method == 'spatial':
            logger.info(f"🔍 Spatial search diagnostics for query: '{query[:50]}...'")

            # Проверяем состояние token_search_engine
            if hasattr(self.token_search_engine, 'faiss_index'):
                logger.info(f"  FAISS index available: {self.token_search_engine.faiss_index is not None}")
                if self.token_search_engine.faiss_index:
                    logger.info(f"  FAISS index vectors: {self.token_search_engine.faiss_index.ntotal}")

            if hasattr(self.token_search_engine, 'embeddings_matrix'):
                logger.info(f"  Embeddings matrix available: {self.token_search_engine.embeddings_matrix is not None}")
                if self.token_search_engine.embeddings_matrix is not None:
                    logger.info(f"  Embeddings matrix shape: {self.token_search_engine.embeddings_matrix.shape}")

            if hasattr(self.token_search_engine, 'vectorizer'):
                logger.info(f"  Vectorizer available: {self.token_search_engine.vectorizer is not None}")
                if self.token_search_engine.vectorizer:
                    logger.info(f"  Vectorizer fitted: {getattr(self.token_search_engine.vectorizer, 'is_fitted', False)}")

            if hasattr(self.token_search_engine, 'tokenizer'):
                logger.info(f"  Tokenizer available: {self.token_search_engine.tokenizer is not None}")

        try:
            results = self.token_search_engine.search_by_tokens(
                query=query,
                method=search_method,
                top_k=top_k
            )

            logger.info(f"Token search ({search_method}) found {len(results)} results for query: {query[:50]}...")

            # Дополнительная диагностика для spatial поиска
            if search_method == 'spatial' and len(results) == 0:
                logger.warning(f"⚠️ Spatial search returned 0 results for query: '{query[:50]}...'")
                logger.warning("  This might indicate issues with:")
                logger.warning("  - FAISS index not properly initialized")
                logger.warning("  - Embeddings matrix is empty or None")
                logger.warning("  - Vectorizer not fitted or unavailable")
                logger.warning("  - Query embedding generation failed")

            return results

        except Exception as e:
            logger.error(f"Error in token search: {e}")
            import traceback
            traceback.print_exc()
            return []

    def batch_search_by_tokens(self, queries: List[str], method: str = None, top_k: int = 10) -> Dict[str, List[SearchResult]]:
        """
        Пакетный поиск аналогов с использованием системы поиска по токенам

        Args:
            queries: Список поисковых запросов
            method: Метод поиска (если None, используется self.search_method)
            top_k: Количество результатов для каждого запроса

        Returns:
            Словарь {запрос: список результатов}
        """
        if not self.token_search_engine:
            logger.error("Token search engine not initialized")
            return {}

        search_method = method or self.search_method
        results = {}

        logger.info(f"Starting batch token search for {len(queries)} queries using method: {search_method}")

        for i, query in enumerate(queries):
            try:
                query_results = self.token_search_engine.search_by_tokens(
                    query=query,
                    method=search_method,
                    top_k=top_k
                )
                results[query] = query_results

                if (i + 1) % 100 == 0:
                    logger.info(f"Processed {i + 1}/{len(queries)} queries")

            except Exception as e:
                logger.error(f"Error in batch token search for query '{query}': {e}")
                results[query] = []

        logger.info(f"Batch token search completed. Processed {len(queries)} queries")
        return results

    def validate_input_csv(self, csv_path: str) -> bool:
        """
        Проверка входного CSV файла на наличие обязательных колонок

        Поддерживает два формата CSV файлов:
        1. Базовый формат (legacy): ['Код', 'Raw_Name', 'Cleaned_Name', 'Lemmatized_Name', 'Normalized_Name', 'parameters', 'Dublikat']
        2. Расширенный формат (advanced): все поля из excel_processor_advanced.py

        Args:
            csv_path: Путь к CSV файлу

        Returns:
            True если файл валиден, False иначе
        """
        # Базовые обязательные колонки (минимум для работы)
        basic_required_columns = ['Код', 'Raw_Name', 'Normalized_Name']

        # Полный список колонок расширенного формата
        advanced_columns = [
            'Код', 'Наименование', 'НаименованиеПолное', 'Группа', 'ВидНоменклатуры', 'ЕдиницаИзмерения',
            'Raw_Name', 'Cleaned_Name', 'Lemmatized_Name', 'Normalized_Name', 'BPE_Tokens', 'BPE_Tokens_Count',
            'Semantic_Category', 'Technical_Complexity', 'Parameter_Confidence', 'Embedding_Similarity',
            'Advanced_Parameters', 'ML_Validated_Parameters', 'Colors_Found', 'Colors_Count',
            'Technical_Terms_Found', 'Technical_Terms_Count', 'Dublikat', 'Processing_Status'
        ]

        # Базовые колонки legacy формата
        legacy_columns = ['Код', 'Raw_Name', 'Cleaned_Name', 'Lemmatized_Name',
                         'Normalized_Name', 'parameters', 'Dublikat']

        try:
            df = pd.read_csv(csv_path)
            available_columns = set(df.columns)

            # Проверяем базовые обязательные колонки
            missing_basic = [col for col in basic_required_columns if col not in available_columns]
            if missing_basic:
                logger.error(f"Отсутствуют базовые обязательные колонки: {missing_basic}")
                return False

            # Определяем формат файла
            advanced_columns_present = sum(1 for col in advanced_columns if col in available_columns)
            legacy_columns_present = sum(1 for col in legacy_columns if col in available_columns)

            if advanced_columns_present >= len(advanced_columns) * 0.8:  # 80% колонок расширенного формата
                self.input_format = 'advanced'
                logger.info(f"Обнаружен расширенный формат CSV файла ({advanced_columns_present}/{len(advanced_columns)} колонок)")
            elif legacy_columns_present >= len(legacy_columns) * 0.8:  # 80% колонок legacy формата
                self.input_format = 'legacy'
                logger.info(f"Обнаружен базовый формат CSV файла ({legacy_columns_present}/{len(legacy_columns)} колонок)")
            else:
                self.input_format = 'basic'
                logger.info("Обнаружен минимальный формат CSV файла (только базовые колонки)")

            logger.info(f"CSV файл валиден. Найдено {len(df)} записей, формат: {self.input_format}")
            return True

        except Exception as e:
            logger.error(f"Ошибка при чтении CSV файла: {e}")
            return False
    
    async def initialize_search_engine(self, catalog_df: pd.DataFrame):
        """
        Инициализация поискового движка с каталогом данных

        Args:
            catalog_df: DataFrame с каталогом товаров
        """
        logger.info("Инициализация поискового движка...")

        # Инициализация новой системы расширенного поиска
        if self.use_extended_search and self.extended_search_interface:
            try:
                logger.info("Initializing extended search interface...")
                # Создаем временный CSV файл для инициализации extended search
                temp_csv_for_extended = self._prepare_token_search_data(catalog_df)
                success = self.extended_search_interface.initialize(temp_csv_for_extended)

                # Удаляем временный файл
                if Path(temp_csv_for_extended).exists():
                    Path(temp_csv_for_extended).unlink()

                if success:
                    logger.info("✅ Extended search interface initialized successfully")
                    # Проверяем доступность компонентов
                    if (hasattr(self.extended_search_interface, 'processor') and
                        self.extended_search_interface.processor and
                        hasattr(self.extended_search_interface.processor, 'tokenizer')):
                        tokenizer = self.extended_search_interface.processor.tokenizer
                        if hasattr(tokenizer, '_vectorizer') and tokenizer._vectorizer:
                            logger.info("✅ Tokenizer and vectorizer available in extended search")
                        else:
                            logger.warning("⚠️ Vectorizer not available in extended search tokenizer")
                    else:
                        logger.warning("⚠️ Processor or tokenizer not available in extended search")
                else:
                    logger.warning("❌ Extended search interface initialization failed, falling back to legacy search")
                    self.use_extended_search = False
            except Exception as e:
                logger.error(f"Error initializing extended search: {e}")
                import traceback
                traceback.print_exc()
                self.use_extended_search = False

        # Инициализация системы поиска по токенам
        if self.use_token_search and self.token_search_engine:
            try:
                logger.info("Initializing token search engine...")
                # Создаем временный CSV файл для инициализации
                temp_csv_path = self._prepare_token_search_data(catalog_df)

                # Получаем токенизатор и векторизатор из расширенного поиска
                tokenizer = None
                vectorizer = None

                if self.extended_search_interface and hasattr(self.extended_search_interface, 'processor'):
                    processor = self.extended_search_interface.processor
                    if processor and hasattr(processor, 'tokenizer'):
                        tokenizer = processor.tokenizer
                        logger.info("✅ Tokenizer obtained from extended search interface")
                        if hasattr(tokenizer, '_vectorizer'):
                            vectorizer = tokenizer._vectorizer
                            if vectorizer and hasattr(vectorizer, 'is_fitted') and vectorizer.is_fitted:
                                logger.info("✅ Vectorizer obtained and is fitted")
                            else:
                                logger.warning("⚠️ Vectorizer not fitted or unavailable")
                        else:
                            logger.warning("⚠️ Vectorizer not found in tokenizer")
                    else:
                        logger.warning("⚠️ Tokenizer not found in processor")
                else:
                    logger.warning("⚠️ Extended search interface or processor not available")

                # Загружаем данные в движок поиска по токенам с токенизатором и векторизатором
                logger.info(f"Loading token search data with tokenizer: {tokenizer is not None}, vectorizer: {vectorizer is not None}")
                success = self.token_search_engine.load_data(temp_csv_path, vectorizer, tokenizer)
                if success:
                    logger.info("✅ Token search engine initialized successfully")
                else:
                    logger.warning("❌ Token search engine initialization failed")
                    self.use_token_search = False

                # Удаляем временный файл
                if Path(temp_csv_path).exists():
                    Path(temp_csv_path).unlink()

            except Exception as e:
                logger.error(f"Error initializing token search: {e}")
                self.use_token_search = False

        # Для больших датасетов используем подвыборку для инициализации
        if len(catalog_df) > 10000:
            logger.info(f"Большой датасет ({len(catalog_df)} записей), используем подвыборку для инициализации")
            # Берем случайную выборку для инициализации поисковых движков
            sample_size = min(5000, len(catalog_df))
            catalog_sample = catalog_df.sample(n=sample_size, random_state=42)
            logger.info(f"Используем выборку из {sample_size} записей для инициализации")
        else:
            catalog_sample = catalog_df

        # Инициализация Hybrid DBSCAN поискового движка
        if self.use_hybrid_dbscan:
            try:
                logger.info("Initializing Hybrid DBSCAN search engine...")
                dbscan_config = HybridDBSCANConfig(
                    eps=0.3,
                    min_samples=2,
                    similarity_threshold=self.similarity_threshold,
                    max_features=10000,
                    batch_size=min(1000, len(catalog_df) // 10)
                )
                self.hybrid_dbscan_engine = HybridDBSCANSearchEngine(dbscan_config)
                logger.info("✅ Hybrid DBSCAN search engine initialized")
            except Exception as e:
                logger.error(f"Error initializing Hybrid DBSCAN engine: {e}")
                self.use_hybrid_dbscan = False
                self.hybrid_dbscan_engine = None

        # Создаем поисковый движок
        self.search_engine = AnalogSearchEngine(self.search_config)

        # Подготавливаем данные для инициализации
        # Используем Normalized_Name как основной текст для поиска
        catalog_for_search = catalog_sample.copy()
        catalog_for_search['name'] = catalog_sample['Normalized_Name'].fillna(catalog_sample['Raw_Name'])

        try:
            # Инициализируем движок с подвыборкой
            await self.search_engine.initialize(catalog_for_search)
            logger.info("Поисковый движок инициализирован")

            # Сохраняем полный каталог для поиска
            self.full_catalog = catalog_df

        except Exception as e:
            logger.error(f"Ошибка инициализации поискового движка: {e}")
            # Пробуем с еще меньшей выборкой
            if len(catalog_sample) > 1000:
                logger.info("Пробуем с меньшей выборкой (1000 записей)")
                smaller_sample = catalog_df.sample(n=1000, random_state=42)
                catalog_for_search = smaller_sample.copy()
                catalog_for_search['name'] = smaller_sample['Normalized_Name'].fillna(smaller_sample['Raw_Name'])
                await self.search_engine.initialize(catalog_for_search)
                self.full_catalog = catalog_df
                logger.info("Поисковый движок инициализирован с меньшей выборкой")
            else:
                raise

    def _prepare_token_search_data(self, catalog_df: pd.DataFrame) -> str:
        """
        Подготавливает данные для системы поиска по токенам

        Args:
            catalog_df: DataFrame с каталогом

        Returns:
            Путь к временному CSV файлу
        """
        try:
            # Создаем временный DataFrame с необходимыми колонками
            token_search_df = catalog_df.copy()

            # Убеждаемся, что есть все необходимые колонки
            required_columns = ['Код', 'Raw_Name']
            for col in required_columns:
                if col not in token_search_df.columns:
                    token_search_df[col] = ''

            # Обрабатываем параметры в зависимости от формата файла
            if self.input_format == 'advanced':
                # Для расширенного формата используем Advanced_Parameters
                if 'Advanced_Parameters' in token_search_df.columns:
                    token_search_df['parameters'] = token_search_df['Advanced_Parameters'].fillna('')
                else:
                    token_search_df['parameters'] = ''
            elif self.input_format == 'legacy':
                # Для legacy формата используем существующую колонку parameters
                if 'parameters' not in token_search_df.columns:
                    token_search_df['parameters'] = ''
            else:
                # Для базового формата создаем пустую колонку parameters
                token_search_df['parameters'] = ''

            # Добавляем колонки для токенов если их нет
            if 'tokenizer' not in token_search_df.columns:
                # Используем BPE_Tokens если доступны (расширенный формат)
                if self.input_format == 'advanced' and 'BPE_Tokens' in token_search_df.columns:
                    # Преобразуем BPE токены из строкового представления в список
                    def parse_bpe_tokens(tokens_str):
                        if pd.isna(tokens_str) or tokens_str == '':
                            return []
                        try:
                            # Убираем квадратные скобки и разделяем по запятым
                            tokens_str = str(tokens_str).strip("[]'\"")
                            if tokens_str:
                                return [token.strip("'\" ") for token in tokens_str.split(',')]
                            return []
                        except:
                            return []

                    token_search_df['tokenizer'] = token_search_df['BPE_Tokens'].apply(parse_bpe_tokens)
                else:
                    # Простая токенизация как fallback
                    token_search_df['tokenizer'] = token_search_df['Normalized_Name'].fillna(
                        token_search_df['Raw_Name']
                    ).str.lower().str.split()

            if 'token_vectors' not in token_search_df.columns:
                # Заглушка для векторов токенов
                token_search_df['token_vectors'] = ''

            # Сохраняем во временный файл
            temp_path = "temp_token_search_data.csv"
            token_search_df.to_csv(temp_path, index=False)

            logger.info(f"Prepared token search data: {len(token_search_df)} records, format: {self.input_format}")
            return temp_path

        except Exception as e:
            logger.error(f"Error preparing token search data: {e}")
            raise

    async def process_catalog(self, input_csv_path: str, output_excel_path: str = None, limit_records: int = None) -> str:
        """
        Основная функция обработки каталога
        
        Args:
            input_csv_path: Путь к входному CSV файлу
            output_excel_path: Путь к выходному Excel файлу (опционально)
            limit_records: Ограничить количество обрабатываемых записей (опционально)

        Returns:
            Путь к созданному Excel файлу
        """
        logger.info(f"Начало обработки каталога: {input_csv_path}")
        
        # Валидация входного файла
        if not self.validate_input_csv(input_csv_path):
            raise ValueError("Входной CSV файл не прошел валидацию")
        
        # Загрузка данных
        catalog_df = pd.read_csv(input_csv_path)
        logger.info(f"Загружено {len(catalog_df)} записей из каталога")

        # Применяем ограничение если указано
        if limit_records and limit_records < len(catalog_df):
            catalog_df = catalog_df.head(limit_records)
            logger.info(f"Ограничено до {limit_records} записей для обработки")
        
        # Специальная обработка для Optimized DBSCAN
        if self.use_optimized_dbscan:
            logger.info("Using Optimized DBSCAN search engine for processing...")

            # Инициализируем Optimized DBSCAN движок если еще не инициализирован
            if not self.optimized_dbscan_engine:
                try:
                    logger.info("Initializing Optimized DBSCAN search engine...")
                    optimized_config = OptimizedDBSCANConfig(
                        eps=0.4,
                        min_samples=3,
                        similarity_threshold=self.similarity_threshold,
                        max_features=5000,
                        batch_size=max(1, min(500, len(catalog_df) // 20)),
                        use_sampling=True,
                        sample_size=min(5000, len(catalog_df)),
                        memory_limit_gb=4.0
                    )
                    self.optimized_dbscan_engine = OptimizedHybridDBSCANSearchEngine(optimized_config)
                    self.optimized_dbscan_engine.set_catalog(catalog_df)
                    logger.info("✅ Optimized DBSCAN search engine initialized")
                except Exception as e:
                    logger.error(f"Error initializing Optimized DBSCAN engine: {e}")
                    # Fallback к обычному поиску
                    self.use_optimized_dbscan = False
                    self.optimized_dbscan_engine = None

            if self.optimized_dbscan_engine:
                # Определяем выходной файл для CSV
                if output_excel_path:
                    csv_output_path = output_excel_path.replace('.xlsx', '.csv').replace('.xls', '.csv')
                else:
                    timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
                    csv_output_path = f"optimized_dbscan_results_{timestamp}.csv"

                # Обрабатываем каталог с помощью Optimized DBSCAN
                logger.info("Starting optimized processing...")

                # Простая обработка без полной кластеризации
                all_results = []
                batch_size = self.optimized_dbscan_engine.config.batch_size

                for i in range(0, len(catalog_df), batch_size):
                    batch_end = min(i + batch_size, len(catalog_df))
                    logger.info(f"Processing batch: items {i}-{batch_end-1}")

                    batch_tasks = []
                    for idx in range(i, batch_end):
                        batch_tasks.append(self.optimized_dbscan_engine.search_analogs_optimized(idx, max_results=5))

                    # Выполняем поиск для батча
                    batch_results = await asyncio.gather(*batch_tasks, return_exceptions=True)

                    for idx, results in enumerate(batch_results):
                        if isinstance(results, Exception):
                            logger.error(f"Error processing item {i + idx}: {results}")
                            continue

                        all_results.extend(results)

                # Сохраняем результаты
                if all_results:
                    results_df = pd.DataFrame(all_results)
                    results_df.to_csv(csv_output_path, index=False, encoding='utf-8')
                    logger.info(f"Results saved to {csv_output_path}")

                logger.info(f"Optimized DBSCAN processing completed. Results saved to: {csv_output_path}")
                return csv_output_path

        # Специальная обработка для Hybrid DBSCAN
        elif self.use_hybrid_dbscan:
            logger.info("Using Hybrid DBSCAN search engine for processing...")

            # Инициализируем Hybrid DBSCAN движок если еще не инициализирован
            if not self.hybrid_dbscan_engine:
                try:
                    logger.info("Initializing Hybrid DBSCAN search engine...")
                    dbscan_config = HybridDBSCANConfig(
                        eps=0.3,
                        min_samples=2,
                        similarity_threshold=self.similarity_threshold,
                        max_features=10000,
                        batch_size=max(1, min(1000, len(catalog_df)))  # Убеждаемся что batch_size >= 1
                    )
                    self.hybrid_dbscan_engine = HybridDBSCANSearchEngine(dbscan_config)
                    logger.info("✅ Hybrid DBSCAN search engine initialized")
                except Exception as e:
                    logger.error(f"Error initializing Hybrid DBSCAN engine: {e}")
                    # Fallback к обычному поиску
                    self.use_hybrid_dbscan = False
                    self.hybrid_dbscan_engine = None

            if self.hybrid_dbscan_engine:
                # Определяем выходной файл для CSV
                if output_excel_path:
                    csv_output_path = output_excel_path.replace('.xlsx', '.csv').replace('.xls', '.csv')
                else:
                    timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
                    csv_output_path = f"hybrid_dbscan_results_{timestamp}.csv"

                # Обрабатываем каталог с помощью Hybrid DBSCAN
                results_df = await self.hybrid_dbscan_engine.process_catalog(catalog_df, csv_output_path)

                logger.info(f"Hybrid DBSCAN processing completed. Results saved to: {csv_output_path}")
                return csv_output_path

        # Инициализация поискового движка
        await self.initialize_search_engine(catalog_df)

        # Обработка каждой записи для поиска аналогов
        results = []
        total_items = len(catalog_df)

        logger.info(f"Начало поиска аналогов для {total_items} записей...")

        # Для больших датасетов обрабатываем батчами
        batch_size = 100 if total_items > 10000 else 10
        processed_count = 0

        for start_idx in range(0, total_items, batch_size):
            end_idx = min(start_idx + batch_size, total_items)
            batch_df = catalog_df.iloc[start_idx:end_idx]

            logger.info(f"Обработка батча {start_idx}-{end_idx} ({processed_count}/{total_items} записей, {processed_count/total_items*100:.1f}%)")

            # Обрабатываем батч
            batch_results = await self._process_batch(batch_df, catalog_df)
            results.extend(batch_results)

            processed_count += len(batch_df)

            # Принудительная очистка памяти каждые 1000 записей
            if processed_count % 1000 == 0:
                import gc
                gc.collect()
                logger.info(f"Очистка памяти после {processed_count} записей")
        
        logger.info(f"Поиск аналогов завершен. Найдено {len(results)} результатов")
        
        # Создание выходного Excel файла
        if not output_excel_path:
            timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
            output_excel_path = f"analog_search_results_{timestamp}.xlsx"
        
        excel_path = await self._create_excel_output(results, output_excel_path)
        
        logger.info(f"Результаты сохранены в: {excel_path}")
        return excel_path

    async def _process_batch(self, batch_df: pd.DataFrame, full_catalog_df: pd.DataFrame) -> List[Dict[str, Any]]:
        """
        Обработка батча записей для поиска аналогов

        Args:
            batch_df: Батч записей для обработки
            full_catalog_df: Полный каталог для исключения самих записей из результатов

        Returns:
            Список результатов поиска для батча
        """
        batch_results = []

        for idx, row in batch_df.iterrows():
            try:
                # Поиск аналогов для текущей записи
                item_results = await self._process_single_item(row, full_catalog_df)
                batch_results.extend(item_results)

            except Exception as e:
                logger.error(f"Ошибка при обработке записи {idx}: {e}")
                # Добавляем запись об ошибке
                error_result = self._create_error_result(row, str(e))
                batch_results.append(error_result)

        return batch_results
    
    async def _process_single_item(self, row: pd.Series, catalog_df: pd.DataFrame) -> List[Dict[str, Any]]:
        """
        Обработка одной записи каталога для поиска аналогов

        Args:
            row: Строка каталога для обработки
            catalog_df: Полный каталог для исключения самой записи из результатов

        Returns:
            Список результатов поиска для данной записи
        """
        # Используем нормализованное наименование для поиска
        search_query = row['Normalized_Name'] if pd.notna(row['Normalized_Name']) else row['Raw_Name']

        if not search_query or search_query.strip() == '':
            return [self._create_error_result(row, "Пустой поисковый запрос")]

        try:
            # Проверяем, можем ли использовать систему поиска по токенам
            if self.use_token_search and self.token_search_engine:
                return await self._process_with_token_search(row, search_query, catalog_df)

            # Проверяем, можем ли использовать новую систему расширенного поиска
            elif self.use_extended_search and self.extended_search_interface and self.extended_search_interface.is_ready:
                # Специальная обработка для поиска по ID токенов
                if self.search_method == "token_id":
                    return await self._process_with_token_id_search(row, search_query)
                else:
                    return await self._process_with_extended_search(row, search_query, catalog_df)

            # Fallback к старой системе поиска
            return await self._process_with_legacy_search(row, search_query, catalog_df)

        except Exception as e:
            logger.error(f"Ошибка поиска аналогов для {row.get('Код', 'unknown')}: {e}")
            return [self._create_error_result(row, str(e))]

    async def _process_with_token_search(self, row: pd.Series, search_query: str, catalog_df: pd.DataFrame) -> List[Dict[str, Any]]:
        """
        Обработка записи с использованием системы поиска по токенам

        Args:
            row: Строка каталога для обработки
            search_query: Поисковый запрос
            catalog_df: Полный каталог

        Returns:
            Список результатов поиска
        """
        try:
            # Выполняем поиск с использованием системы поиска по токенам
            search_results = self.search_by_tokens(
                query=search_query,
                method=self.search_method,
                top_k=15  # Больше результатов для лучшей фильтрации
            )

            # Обрабатываем результаты
            processed_results = []
            original_code = row['Код']

            for result in search_results:
                # Исключаем саму запись из результатов
                if result.code == original_code:
                    continue

                # Адаптивный порог для spatial поиска (TF-IDF эмбеддинги дают низкие оценки)
                effective_threshold = self.similarity_threshold
                if self.search_method == 'spatial':
                    effective_threshold = max(0.05, self.similarity_threshold * 0.3)  # Снижаем порог для spatial

                # Проверяем порог схожести
                if result.score < effective_threshold:
                    continue

                # Создаем результат в требуемом формате
                processed_result = self._format_token_search_result(row, result)
                processed_results.append(processed_result)

            # Если не найдено аналогов, создаем запись об отсутствии результатов
            if not processed_results:
                no_results = self._create_no_results_entry(row)
                no_results['Search_Engine'] = f"token_search_{self.search_method}"
                no_results['Comment'] = f"Поиск по токенам ({self.search_method}) не дал результатов"
                processed_results.append(no_results)

            logger.debug(f"Token search ({self.search_method}) found {len(processed_results)} results for {original_code}")
            return processed_results

        except Exception as e:
            logger.error(f"Error in token search for {row.get('Код', 'unknown')}: {e}")
            # Fallback к расширенному поиску если доступен
            if self.use_extended_search and self.extended_search_interface and self.extended_search_interface.is_ready:
                return await self._process_with_extended_search(row, search_query, catalog_df)
            # Иначе fallback к старой системе
            return await self._process_with_legacy_search(row, search_query, catalog_df)

    async def _process_with_extended_search(self, row: pd.Series, search_query: str, catalog_df: pd.DataFrame) -> List[Dict[str, Any]]:
        """
        Обработка записи с использованием новой системы расширенного поиска

        Args:
            row: Строка каталога для обработки
            search_query: Поисковый запрос
            catalog_df: Полный каталог

        Returns:
            Список результатов поиска
        """
        try:
            # Определяем метод поиска
            if self.search_method == "token_id":
                search_method = "token_id"
            elif self.search_method in ["hybrid", "extended_hybrid"]:
                search_method = "extended_hybrid"
            else:
                search_method = self.search_method

            # Выполняем поиск с новой системой
            search_results = self.extended_search_interface.search(
                search_query,
                method=search_method,
                top_k=10
            )

            # Обрабатываем результаты
            processed_results = []
            original_code = row['Код']

            for result in search_results:
                # Исключаем саму запись из результатов
                if result.code == original_code:
                    continue

                # Проверяем порог схожести
                if result.score < self.similarity_threshold:
                    continue

                # Создаем результат в требуемом формате
                processed_result = self._format_extended_search_result(row, result)
                processed_results.append(processed_result)

            # Если не найдено аналогов, создаем запись об отсутствии результатов
            if not processed_results:
                no_results = self._create_no_results_entry(row)
                no_results['Search_Engine'] = f"extended_{search_method}"
                processed_results.append(no_results)

            logger.debug(f"Extended search found {len(processed_results)} results for {original_code}")
            return processed_results

        except Exception as e:
            logger.error(f"Error in extended search for {row.get('Код', 'unknown')}: {e}")
            # Fallback к старой системе
            return await self._process_with_legacy_search(row, search_query, catalog_df)

    async def _process_with_legacy_search(self, row: pd.Series, search_query: str, catalog_df: pd.DataFrame) -> List[Dict[str, Any]]:
        """
        Обработка записи с использованием старой системы поиска

        Args:
            row: Строка каталога для обработки
            search_query: Поисковый запрос
            catalog_df: Полный каталог

        Returns:
            Список результатов поиска
        """
        # Выполняем поиск аналогов
        if self.search_engine is None:
            logger.warning("Поисковый движок не инициализирован, используем простой поиск")
            return [self._create_simple_search_result(row, catalog_df)]

        search_results = await self.search_engine.search_analogs_async([search_query], self.search_method)
        query_results = search_results.get(search_query, [])

        # Обрабатываем результаты
        processed_results = []
        original_code = row['Код']

        for result in query_results:
            # Исключаем саму запись из результатов (по коду)
            candidate_row_idx = result.get('document_id')
            if candidate_row_idx is not None and candidate_row_idx < len(catalog_df):
                candidate_row = catalog_df.iloc[candidate_row_idx]
                candidate_code = candidate_row.get('Код', '')

                # Пропускаем если это та же самая запись
                if candidate_code == original_code:
                    continue

            # Создаем результат в требуемом формате
            processed_result = self._format_search_result(row, result, catalog_df)
            processed_results.append(processed_result)

        # Если не найдено аналогов, создаем запись об отсутствии результатов
        if not processed_results:
            no_results = self._create_no_results_entry(row)
            processed_results.append(no_results)

        return processed_results

    async def _process_with_token_id_search(self, row: pd.Series, search_query: str) -> List[Dict[str, Any]]:
        """
        Обработка записи с использованием чистого поиска по ID токенов

        Args:
            row: Строка каталога для обработки
            search_query: Поисковый запрос

        Returns:
            Список результатов поиска по ID токенов
        """
        try:
            # Выполняем поиск только по ID токенов
            search_results = self.extended_search_interface.search(
                search_query,
                method="token_id",
                top_k=15  # Больше результатов для поиска по токенам
            )

            # Обрабатываем результаты
            processed_results = []
            original_code = row['Код']

            for result in search_results:
                # Исключаем саму запись из результатов
                if result.code == original_code:
                    continue

                # Проверяем порог схожести (для поиска по токенам можем использовать более низкий порог)
                token_id_threshold = max(0.1, self.similarity_threshold * 0.5)  # Половина от основного порога
                if result.score < token_id_threshold:
                    continue

                # Создаем результат в требуемом формате
                processed_result = self._format_token_id_search_result(row, result)
                processed_results.append(processed_result)

            # Если не найдено аналогов, создаем запись об отсутствии результатов
            if not processed_results:
                no_results = self._create_no_results_entry(row)
                no_results['Search_Engine'] = "token_id_search"
                no_results['Comment'] = "Поиск по ID токенов не дал результатов"
                processed_results.append(no_results)

            logger.debug(f"Token ID search found {len(processed_results)} results for {original_code}")
            return processed_results

        except Exception as e:
            logger.error(f"Error in token ID search for {row.get('Код', 'unknown')}: {e}")
            # Создаем запись об ошибке
            error_result = self._create_error_result(row, f"Ошибка поиска по ID токенов: {str(e)}")
            error_result['Search_Engine'] = "token_id_search_error"
            return [error_result]

    def _format_token_search_result(self, original_row: pd.Series, search_result) -> Dict[str, Any]:
        """
        Форматирование результата поиска по токенам

        Args:
            original_row: Исходная запись каталога
            search_result: Результат поиска по токенам (SearchResult)

        Returns:
            Отформатированный результат
        """
        # Определяем категорию оригинального товара
        if self.input_format == 'advanced' and 'Semantic_Category' in original_row and pd.notna(original_row['Semantic_Category']):
            original_category = original_row['Semantic_Category']
        else:
            original_category, _ = self.category_classifier.classify(original_row['Raw_Name'])

        # Определяем тип связи на основе схожести
        similarity_score = search_result.score
        relation_type = self._determine_relation_type(similarity_score)

        # Создаем детальную информацию о методе поиска
        search_details = []
        method_name = {
            'token_id': 'Поиск по ID токенов',
            'prefix': 'Префиксный поиск',
            'inverted_index': 'Инвертированный индекс',
            'tfidf': 'TF-IDF поиск',
            'lsh': 'LSH поиск',
            'spatial': 'Пространственный поиск',
            'advanced_hybrid': 'Продвинутый гибридный поиск'
        }.get(search_result.match_type, search_result.match_type)

        search_details.append(f"Метод: {method_name}")

        # Добавляем информацию о совпадающих токенах
        if hasattr(search_result, 'matched_tokens') and search_result.matched_tokens:
            try:
                if isinstance(search_result.matched_tokens, list):
                    tokens_str = ', '.join(str(token) for token in search_result.matched_tokens[:5])  # Первые 5 токенов
                    if len(search_result.matched_tokens) > 5:
                        tokens_str += f" (и еще {len(search_result.matched_tokens) - 5})"
                    search_details.append(f"Совпадающие токены: {tokens_str}")
                else:
                    search_details.append(f"Совпадающие токены: {search_result.matched_tokens}")
            except Exception:
                search_details.append("Совпадающие токены: информация недоступна")

        # Добавляем информацию о семантическом сходстве если доступно
        if hasattr(search_result, 'similarity_score') and search_result.similarity_score > 0:
            search_details.append(f"Семантическое сходство: {search_result.similarity_score:.3f}")

        # Добавляем информацию из расширенного формата если доступна
        if self.input_format == 'advanced':
            if 'Technical_Complexity' in original_row and pd.notna(original_row['Technical_Complexity']):
                search_details.append(f"Техническая сложность: {original_row['Technical_Complexity']}")
            if 'Colors_Found' in original_row and pd.notna(original_row['Colors_Found']) and original_row['Colors_Found']:
                search_details.append(f"Цвета: {original_row['Colors_Found']}")
            if 'Technical_Terms_Found' in original_row and pd.notna(original_row['Technical_Terms_Found']) and original_row['Technical_Terms_Found']:
                search_details.append(f"Технические термины: {original_row['Technical_Terms_Found']}")

        # Формируем комментарий
        comment = "; ".join(search_details)

        # Создаем базовый результат
        result = {
            'Код': original_row['Код'],
            'Raw_Name': original_row['Raw_Name'],
            'Candidate_Name': search_result.raw_name,
            'Similarity_Score': round(similarity_score, 4),
            'Relation_Type': relation_type,
            'Suggested_Category': original_category,
            'Final_Decision': '',  # Пустое поле для ручного заполнения
            'Comment': comment,
            'Original_Category': original_category,
            'Original_Code': original_row['Код'],
            'Search_Engine': f"token_search_{search_result.match_type}"
        }

        # Добавляем дополнительные поля из расширенного формата если доступны
        if self.input_format == 'advanced':
            additional_fields = {
                'Наименование': original_row.get('Наименование', ''),
                'НаименованиеПолное': original_row.get('НаименованиеПолное', ''),
                'Группа': original_row.get('Группа', ''),
                'ВидНоменклатуры': original_row.get('ВидНоменклатуры', ''),
                'ЕдиницаИзмерения': original_row.get('ЕдиницаИзмерения', ''),
                'Cleaned_Name': original_row.get('Cleaned_Name', ''),
                'Lemmatized_Name': original_row.get('Lemmatized_Name', ''),
                'Normalized_Name': original_row.get('Normalized_Name', ''),
                'BPE_Tokens': original_row.get('BPE_Tokens', ''),
                'BPE_Tokens_Count': original_row.get('BPE_Tokens_Count', ''),
                'Technical_Complexity': original_row.get('Technical_Complexity', ''),
                'Parameter_Confidence': original_row.get('Parameter_Confidence', ''),
                'Embedding_Similarity': original_row.get('Embedding_Similarity', ''),
                'Advanced_Parameters': original_row.get('Advanced_Parameters', ''),
                'ML_Validated_Parameters': original_row.get('ML_Validated_Parameters', ''),
                'Colors_Found': original_row.get('Colors_Found', ''),
                'Colors_Count': original_row.get('Colors_Count', ''),
                'Technical_Terms_Found': original_row.get('Technical_Terms_Found', ''),
                'Technical_Terms_Count': original_row.get('Technical_Terms_Count', ''),
                'Dublikat': original_row.get('Dublikat', ''),
                'Processing_Status': original_row.get('Processing_Status', '')
            }
            result.update(additional_fields)

        return result

    def _format_token_id_search_result(self, original_row: pd.Series, search_result) -> Dict[str, Any]:
        """
        Форматирование результата поиска по ID токенов

        Args:
            original_row: Исходная запись каталога
            search_result: Результат поиска по ID токенов

        Returns:
            Отформатированный результат
        """
        # Определяем категорию оригинального товара
        original_category, _ = self.category_classifier.classify(original_row['Raw_Name'])

        # Определяем тип связи на основе схожести
        similarity_score = search_result.score
        relation_type = self._determine_relation_type(similarity_score)

        # Создаем информацию о совпадающих токенах
        token_info = []
        if hasattr(search_result, 'matched_token_ids') and search_result.matched_token_ids:
            # Преобразуем ID токенов в строки для объединения
            token_ids_str = ', '.join(str(tid) for tid in search_result.matched_token_ids)
            token_info.append(f"Совпадающие ID токенов: {token_ids_str}")

        if hasattr(search_result, 'matched_tokens') and search_result.matched_tokens:
            # Убеждаемся, что все токены - строки
            tokens_str = ', '.join(str(token) for token in search_result.matched_tokens)
            token_info.append(f"Совпадающие токены: {tokens_str}")

        # Добавляем информацию о векторном расстоянии
        if hasattr(search_result, 'vector_distance'):
            token_info.append(f"Векторное расстояние: {search_result.vector_distance:.4f}")

        # Формируем комментарий
        comment_parts = ["Поиск по ID токенов"]
        if token_info:
            comment_parts.extend(token_info)

        comment = "; ".join(comment_parts)

        return {
            'Код': original_row['Код'],
            'Raw_Name': original_row['Raw_Name'],
            'Candidate_Name': search_result.raw_name,
            'Similarity_Score': round(similarity_score, 4),
            'Relation_Type': relation_type,
            'Suggested_Category': original_category,
            'Final_Decision': '',  # Пустое поле для ручного заполнения
            'Comment': comment,  # Информация о токенах
            'Original_Category': original_category,
            'Original_Code': original_row['Код'],
            'Search_Engine': "token_id_search"
        }

    def _format_extended_search_result(self, original_row: pd.Series, search_result) -> Dict[str, Any]:
        """
        Форматирование результата расширенного поиска в требуемый формат

        Args:
            original_row: Исходная запись каталога
            search_result: Результат поиска от расширенной системы (SearchResult)

        Returns:
            Отформатированный результат
        """
        # Определяем категорию оригинального товара
        original_category, _ = self.category_classifier.classify(original_row['Raw_Name'])

        # Определяем тип связи на основе схожести
        similarity_score = search_result.score
        relation_type = self._determine_relation_type(similarity_score)

        # Создаем детальную информацию о методах поиска
        search_details = []
        if hasattr(search_result, 'method_scores') and search_result.method_scores:
            for method, score in search_result.method_scores.items():
                if score > 0:
                    method_name = {
                        'exact_match': 'Точное совпадение',
                        'partial_match': 'Частичное совпадение',
                        'semantic': 'Семантика',
                        'subset_match': 'Подмножества',
                        'token_id': 'Поиск по ID токенов'
                    }.get(method, method)
                    search_details.append(f"{method_name}: {score:.3f}")

        # Специальная обработка для поиска по ID токенов
        if hasattr(search_result, 'match_type') and search_result.match_type == 'token_id':
            if hasattr(search_result, 'matched_token_ids') and search_result.matched_token_ids:
                # Безопасное преобразование ID токенов в строки
                token_ids_str = ', '.join(str(tid) for tid in search_result.matched_token_ids)
                token_ids_info = f"ID токенов: {token_ids_str}"
                search_details.append(token_ids_info)

        # Добавляем информацию о техническом бонусе
        technical_info = ""
        if hasattr(search_result, 'technical_boost') and search_result.technical_boost > 0:
            technical_info = f"Технический бонус: {search_result.technical_boost:.3f}"

        # Формируем комментарий с детальной аналитикой
        comment_parts = []
        if search_details:
            comment_parts.append("Детали: " + "; ".join(search_details))
        if technical_info:
            comment_parts.append(technical_info)
        if hasattr(search_result, 'matched_tokens') and search_result.matched_tokens:
            # Безопасное подсчитывание токенов
            try:
                token_count = len(search_result.matched_tokens)
                comment_parts.append(f"Совпадающих токенов: {token_count}")
            except (TypeError, AttributeError):
                comment_parts.append("Совпадающих токенов: неизвестно")

        comment = "; ".join(comment_parts) if comment_parts else "Расширенный поиск"

        return {
            'Код': original_row['Код'],
            'Raw_Name': original_row['Raw_Name'],
            'Candidate_Name': search_result.raw_name,
            'Similarity_Score': round(similarity_score, 4),
            'Relation_Type': relation_type,
            'Suggested_Category': original_category,
            'Final_Decision': '',  # Пустое поле для ручного заполнения
            'Comment': comment,  # Детальная аналитика
            'Original_Category': original_category,
            'Original_Code': original_row['Код'],
            'Search_Engine': f"extended_{search_result.match_type}"
        }

    def _format_search_result(self, original_row: pd.Series, search_result: Dict[str, Any],
                            catalog_df: pd.DataFrame) -> Dict[str, Any]:
        """
        Форматирование результата поиска в требуемый формат
        
        Args:
            original_row: Исходная запись каталога
            search_result: Результат поиска от движка
            catalog_df: Полный каталог для получения данных кандидата
            
        Returns:
            Отформатированный результат
        """
        # Получаем данные кандидата
        candidate_row_idx = search_result.get('document_id')
        if candidate_row_idx is not None and candidate_row_idx < len(catalog_df):
            candidate_row = catalog_df.iloc[candidate_row_idx]
        else:
            candidate_row = pd.Series()
        
        # Определяем категорию оригинального товара
        original_category, _ = self.category_classifier.classify(original_row['Raw_Name'])
        
        # Определяем тип связи на основе схожести
        similarity_score = search_result.get('similarity_score', search_result.get('combined_score', 0))
        relation_type = self._determine_relation_type(similarity_score)
        
        # Определяем поисковый движок
        search_engine = search_result.get('search_method', self.search_method)
        
        return {
            'Код': original_row['Код'],
            'Raw_Name': original_row['Raw_Name'],
            'Candidate_Name': candidate_row.get('Raw_Name', search_result.get('document', '')),
            'Similarity_Score': round(similarity_score, 4),
            'Relation_Type': relation_type,
            'Suggested_Category': original_category,
            'Final_Decision': '',  # Пустое поле для ручного заполнения
            'Comment': '',  # Пустое поле для комментариев
            'Original_Category': original_category,
            'Original_Code': original_row['Код'],
            'Search_Engine': search_engine
        }

    def _determine_relation_type(self, similarity_score: float) -> str:
        """Определение типа связи на основе оценки схожести"""
        if similarity_score >= 0.9:
            return 'Точный аналог'
        elif similarity_score >= 0.7:
            return 'Близкий аналог'
        elif similarity_score >= 0.5:
            return 'Возможный аналог'
        else:
            return 'Слабое соответствие'

    def _create_error_result(self, row: pd.Series, error_message: str) -> Dict[str, Any]:
        """Создание записи об ошибке"""
        # Создаем базовый результат
        result = {
            'Код': row['Код'],
            'Raw_Name': row['Raw_Name'],
            'Candidate_Name': '',
            'Similarity_Score': 0.0,
            'Relation_Type': 'Ошибка',
            'Suggested_Category': '',
            'Final_Decision': '',
            'Comment': f'Ошибка: {error_message}',
            'Original_Category': '',
            'Original_Code': row['Код'],
            'Search_Engine': 'error'
        }

        # Добавляем дополнительные поля из расширенного формата если доступны
        if self.input_format == 'advanced':
            additional_fields = {
                'Наименование': row.get('Наименование', ''),
                'НаименованиеПолное': row.get('НаименованиеПолное', ''),
                'Группа': row.get('Группа', ''),
                'ВидНоменклатуры': row.get('ВидНоменклатуры', ''),
                'ЕдиницаИзмерения': row.get('ЕдиницаИзмерения', ''),
                'Cleaned_Name': row.get('Cleaned_Name', ''),
                'Lemmatized_Name': row.get('Lemmatized_Name', ''),
                'Normalized_Name': row.get('Normalized_Name', ''),
                'BPE_Tokens': row.get('BPE_Tokens', ''),
                'BPE_Tokens_Count': row.get('BPE_Tokens_Count', ''),
                'Technical_Complexity': row.get('Technical_Complexity', ''),
                'Parameter_Confidence': row.get('Parameter_Confidence', ''),
                'Embedding_Similarity': row.get('Embedding_Similarity', ''),
                'Advanced_Parameters': row.get('Advanced_Parameters', ''),
                'ML_Validated_Parameters': row.get('ML_Validated_Parameters', ''),
                'Colors_Found': row.get('Colors_Found', ''),
                'Colors_Count': row.get('Colors_Count', ''),
                'Technical_Terms_Found': row.get('Technical_Terms_Found', ''),
                'Technical_Terms_Count': row.get('Technical_Terms_Count', ''),
                'Dublikat': row.get('Dublikat', ''),
                'Processing_Status': row.get('Processing_Status', '')
            }
            result.update(additional_fields)

        return result

    def _create_no_results_entry(self, row: pd.Series) -> Dict[str, Any]:
        """Создание записи об отсутствии результатов"""
        # Определяем категорию
        if self.input_format == 'advanced' and 'Semantic_Category' in row and pd.notna(row['Semantic_Category']):
            original_category = row['Semantic_Category']
        else:
            original_category, _ = self.category_classifier.classify(row['Raw_Name'])

        # Создаем базовый результат
        result = {
            'Код': row['Код'],
            'Raw_Name': row['Raw_Name'],
            'Candidate_Name': '',
            'Similarity_Score': 0.0,
            'Relation_Type': 'Аналоги не найдены',
            'Suggested_Category': original_category,
            'Final_Decision': '',
            'Comment': 'Аналоги не найдены в каталоге',
            'Original_Category': original_category,
            'Original_Code': row['Код'],
            'Search_Engine': self.search_method
        }

        # Добавляем дополнительные поля из расширенного формата если доступны
        if self.input_format == 'advanced':
            additional_fields = {
                'Наименование': row.get('Наименование', ''),
                'НаименованиеПолное': row.get('НаименованиеПолное', ''),
                'Группа': row.get('Группа', ''),
                'ВидНоменклатуры': row.get('ВидНоменклатуры', ''),
                'ЕдиницаИзмерения': row.get('ЕдиницаИзмерения', ''),
                'Cleaned_Name': row.get('Cleaned_Name', ''),
                'Lemmatized_Name': row.get('Lemmatized_Name', ''),
                'Normalized_Name': row.get('Normalized_Name', ''),
                'BPE_Tokens': row.get('BPE_Tokens', ''),
                'BPE_Tokens_Count': row.get('BPE_Tokens_Count', ''),
                'Technical_Complexity': row.get('Technical_Complexity', ''),
                'Parameter_Confidence': row.get('Parameter_Confidence', ''),
                'Embedding_Similarity': row.get('Embedding_Similarity', ''),
                'Advanced_Parameters': row.get('Advanced_Parameters', ''),
                'ML_Validated_Parameters': row.get('ML_Validated_Parameters', ''),
                'Colors_Found': row.get('Colors_Found', ''),
                'Colors_Count': row.get('Colors_Count', ''),
                'Technical_Terms_Found': row.get('Technical_Terms_Found', ''),
                'Technical_Terms_Count': row.get('Technical_Terms_Count', ''),
                'Dublikat': row.get('Dublikat', ''),
                'Processing_Status': row.get('Processing_Status', '')
            }
            result.update(additional_fields)

        return result

    def _create_simple_search_result(self, row: pd.Series, catalog_df: pd.DataFrame) -> Dict[str, Any]:
        """
        Создание простого результата поиска без использования поискового движка
        (fallback метод для случаев, когда основной движок недоступен)

        Args:
            row: Исходная запись каталога
            catalog_df: Полный каталог для простого поиска

        Returns:
            Результат простого поиска
        """
        try:
            # Простой поиск по ключевым словам в нормализованном наименовании
            search_query = row['Normalized_Name'] if pd.notna(row['Normalized_Name']) else row['Raw_Name']
            original_code = row['Код']

            # Извлекаем ключевые слова из запроса
            query_words = set(search_query.lower().split())

            best_match = None
            best_score = 0.0

            # Простой поиск по совпадению слов
            for _, candidate_row in catalog_df.iterrows():
                candidate_code = candidate_row.get('Код', '')

                # Пропускаем саму запись
                if candidate_code == original_code:
                    continue

                candidate_name = candidate_row.get('Normalized_Name', candidate_row.get('Raw_Name', ''))
                if not candidate_name:
                    continue

                candidate_words = set(candidate_name.lower().split())

                # Вычисляем простую схожесть по пересечению слов
                intersection = query_words.intersection(candidate_words)
                union = query_words.union(candidate_words)

                if len(union) > 0:
                    score = len(intersection) / len(union)

                    if score > best_score and score >= self.similarity_threshold:
                        best_score = score
                        best_match = candidate_row

            # Определяем категорию
            original_category, _ = self.category_classifier.classify(row['Raw_Name'])

            if best_match is not None:
                return {
                    'Код': original_code,
                    'Raw_Name': row['Raw_Name'],
                    'Candidate_Name': best_match.get('Raw_Name', ''),
                    'Similarity_Score': round(best_score, 4),
                    'Relation_Type': self._determine_relation_type(best_score),
                    'Suggested_Category': original_category,
                    'Final_Decision': '',
                    'Comment': 'Найдено простым поиском (fallback)',
                    'Original_Category': original_category,
                    'Original_Code': original_code,
                    'Search_Engine': 'simple_fallback'
                }
            
            else:
                return {
                    'Код': original_code,
                    'Raw_Name': row['Raw_Name'],
                    'Candidate_Name': '',
                    'Similarity_Score': 0.0,
                    'Relation_Type': 'Аналоги не найдены',
                    'Suggested_Category': original_category,
                    'Final_Decision': '',
                    'Comment': 'Аналоги не найдены простым поиском',
                    'Original_Category': original_category,
                    'Original_Code': original_code,
                    'Search_Engine': 'simple_fallback'
                }

        except Exception as e:
            logger.error(f"Ошибка простого поиска для {row.get('Код', 'unknown')}: {e}")
            return self._create_error_result(row, f"Ошибка простого поиска: {str(e)}")

    async def _create_excel_output(self, results: List[Dict[str, Any]], output_path: str) -> str:
        """
        Создание Excel файла с результатами поиска аналогов

        Args:
            results: Список результатов поиска
            output_path: Путь к выходному файлу

        Returns:
            Путь к созданному файлу
        """
        logger.info(f"Создание Excel файла: {output_path}")

        # Excel имеет ограничение на количество строк (1,048,576)
        # Оставляем место для заголовка, поэтому максимум 1,048,575 строк данных
        EXCEL_MAX_ROWS = 1048575

        # Используем минимальное значение между пользовательским лимитом и лимитом Excel
        effective_max_rows = min(self.max_excel_results, EXCEL_MAX_ROWS)

        total_results = len(results)
        if total_results > effective_max_rows:
            logger.warning(f"Количество результатов ({total_results}) превышает лимит ({effective_max_rows})")

            # Создаем несколько файлов если результатов слишком много
            num_files = (total_results + effective_max_rows - 1) // effective_max_rows  # Округление вверх
            logger.info(f"Результаты будут разбиты на {num_files} файлов")

            created_files = []
            for i in range(num_files):
                start_idx = i * effective_max_rows
                end_idx = min((i + 1) * effective_max_rows, total_results)

                # Создаем имя файла с номером части
                if num_files > 1:
                    base_name = output_path.rsplit('.', 1)[0]
                    extension = output_path.rsplit('.', 1)[1] if '.' in output_path else 'xlsx'
                    part_output_path = f"{base_name}_part{i+1}.{extension}"
                else:
                    part_output_path = output_path

                # Создаем файл для этой части
                part_results = results[start_idx:end_idx]
                part_file = await self._create_single_excel_file(part_results, part_output_path)
                created_files.append(part_file)

                logger.info(f"Создан файл {i+1}/{num_files}: {part_file} ({len(part_results)} записей)")

            # Возвращаем путь к первому файлу
            return created_files[0]
        else:
            # Создаем один файл
            return await self._create_single_excel_file(results, output_path)

    async def _create_single_excel_file(self, results: List[Dict[str, Any]], output_path: str) -> str:
        """
        Создание одного Excel файла с результатами

        Args:
            results: Список результатов поиска
            output_path: Путь к выходному файлу

        Returns:
            Путь к созданному файлу
        """
        # Создаем DataFrame из результатов
        df = pd.DataFrame(results)

        # Создаем Excel файл с форматированием
        wb = Workbook()
        ws = wb.active
        ws.title = "Результаты поиска аналогов"

        # Определяем порядок колонок в зависимости от формата входных данных
        if self.input_format == 'advanced':
            # Расширенный набор колонок для advanced формата
            column_order = [
                # Основные поля поиска аналогов
                'Код', 'Raw_Name', 'Candidate_Name', 'Similarity_Score',
                'Relation_Type', 'Suggested_Category', 'Final_Decision', 'Comment',
                'Original_Category', 'Original_Code', 'Search_Engine',

                # Дополнительные поля из расширенного формата
                'Наименование', 'НаименованиеПолное', 'Группа', 'ВидНоменклатуры', 'ЕдиницаИзмерения',
                'Cleaned_Name', 'Lemmatized_Name', 'Normalized_Name', 'BPE_Tokens', 'BPE_Tokens_Count',
                'Semantic_Category', 'Technical_Complexity', 'Parameter_Confidence', 'Embedding_Similarity',
                'Advanced_Parameters', 'ML_Validated_Parameters', 'Colors_Found', 'Colors_Count',
                'Technical_Terms_Found', 'Technical_Terms_Count', 'Dublikat', 'Processing_Status'
            ]
        else:
            # Базовый набор колонок для legacy и basic форматов
            column_order = [
                'Код', 'Raw_Name', 'Candidate_Name', 'Similarity_Score',
                'Relation_Type', 'Suggested_Category', 'Final_Decision', 'Comment',
                'Original_Category', 'Original_Code', 'Search_Engine'
            ]

        # Фильтруем колонки, которые действительно присутствуют в данных
        available_columns = [col for col in column_order if col in df.columns]

        # Переупорядочиваем DataFrame только по доступным колонкам
        df = df.reindex(columns=available_columns)

        # Записываем данные в лист
        for r in dataframe_to_rows(df, index=False, header=True):
            ws.append(r)

        # Применяем форматирование
        self._apply_excel_formatting(ws, len(df))

        # Сохраняем файл
        wb.save(output_path)

        logger.info(f"Excel файл создан: {output_path}")
        return output_path

    def _apply_excel_formatting(self, worksheet, num_rows: int):
        """
        Применение форматирования к Excel листу

        Args:
            worksheet: Лист Excel для форматирования
            num_rows: Количество строк данных
        """
        # Стиль заголовка
        header_font = Font(bold=True, color="FFFFFF")
        header_fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
        header_alignment = Alignment(horizontal="center", vertical="center")

        # Применяем стили к заголовку
        for cell in worksheet[1]:
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = header_alignment

        # Автоматическая настройка ширины колонок
        for column in worksheet.columns:
            max_length = 0
            column_letter = column[0].column_letter

            for cell in column:
                try:
                    if len(str(cell.value)) > max_length:
                        max_length = len(str(cell.value))
                except:
                    pass

            adjusted_width = min(max_length + 2, 50)  # Максимум 50 символов
            worksheet.column_dimensions[column_letter].width = adjusted_width

        # Добавляем фильтры
        worksheet.auto_filter.ref = worksheet.dimensions

