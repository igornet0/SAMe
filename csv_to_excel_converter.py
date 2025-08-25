#!/usr/bin/env python3
"""
Простой конвертер CSV в Excel с форматированием
"""

import pandas as pd
import argparse
import logging
from pathlib import Path
from datetime import datetime
import os
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils.dataframe import dataframe_to_rows

logging.basicConfig(level=logging.INFO, format='%(asctime)s - %(levelname)s - %(message)s')
logger = logging.getLogger(__name__)


def create_excel_with_formatting(csv_file: str, excel_file: str, metadata: dict = None) -> bool:
    """Создание Excel файла с форматированием из CSV"""
    try:
        logger.info(f"Converting {csv_file} to {excel_file}...")
        
        # Загружаем CSV данные
        df = pd.read_csv(csv_file)
        logger.info(f"Loaded {len(df)} records from CSV")
        
        # Создаем новую книгу Excel
        wb = Workbook()
        
        # Удаляем стандартный лист
        if 'Sheet' in wb.sheetnames:
            wb.remove(wb['Sheet'])
        
        # === ОСНОВНЫЕ ДАННЫЕ (разбиваем на листы если много данных) ===
        max_rows_per_sheet = 1000000  # Оставляем запас для Excel
        total_rows = len(df)

        if total_rows <= max_rows_per_sheet:
            # Один лист
            ws_data = wb.create_sheet("Результаты поиска")

            # Записываем данные
            for r in dataframe_to_rows(df, index=False, header=True):
                ws_data.append(r)

            data_sheets = [ws_data]
        else:
            # Несколько листов
            data_sheets = []
            num_sheets = (total_rows + max_rows_per_sheet - 1) // max_rows_per_sheet

            logger.info(f"Data too large ({total_rows:,} rows), splitting into {num_sheets} sheets")

            for sheet_num in range(num_sheets):
                start_idx = sheet_num * max_rows_per_sheet
                end_idx = min(start_idx + max_rows_per_sheet, total_rows)

                sheet_name = f"Результаты поиска {sheet_num + 1}"
                ws_data = wb.create_sheet(sheet_name)

                # Получаем часть данных
                df_chunk = df.iloc[start_idx:end_idx]

                # Записываем данные
                for r in dataframe_to_rows(df_chunk, index=False, header=True):
                    ws_data.append(r)

                data_sheets.append(ws_data)
                logger.info(f"Created sheet {sheet_num + 1}: rows {start_idx:,}-{end_idx:,}")
        
        # Форматирование всех листов с данными
        header_font = Font(bold=True, color="FFFFFF")
        header_fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
        header_alignment = Alignment(horizontal="center", vertical="center")
        high_similarity_fill = PatternFill(start_color="90EE90", end_color="90EE90", fill_type="solid")

        for ws_data in data_sheets:
            # Форматирование заголовков
            for col in range(1, len(df.columns) + 1):
                cell = ws_data.cell(row=1, column=col)
                cell.font = header_font
                cell.fill = header_fill
                cell.alignment = header_alignment

            # Автоподбор ширины колонок (только для первого листа, чтобы ускорить)
            if ws_data == data_sheets[0]:
                for column in ws_data.columns:
                    max_length = 0
                    column_letter = column[0].column_letter

                    # Проверяем только первые 100 строк для скорости
                    for cell in list(column)[:100]:
                        try:
                            if len(str(cell.value)) > max_length:
                                max_length = len(str(cell.value))
                        except:
                            pass

                    adjusted_width = min(max_length + 2, 50)  # Максимум 50 символов
                    ws_data.column_dimensions[column_letter].width = adjusted_width

                    # Применяем ту же ширину ко всем листам
                    for other_ws in data_sheets[1:]:
                        other_ws.column_dimensions[column_letter].width = adjusted_width

            # Добавляем фильтры
            ws_data.auto_filter.ref = ws_data.dimensions

            # Подсветка высокой схожести (только для небольших листов)
            if 'Similarity_Score' in df.columns and ws_data.max_row <= 10000:
                similarity_col = df.columns.get_loc('Similarity_Score') + 1

                for row in range(2, ws_data.max_row + 1):  # Начинаем с 2-й строки
                    cell = ws_data.cell(row=row, column=similarity_col)
                    try:
                        if cell.value and float(cell.value) >= 0.9:
                            cell.fill = high_similarity_fill
                    except:
                        pass
        
        # === ЛИСТ 2: СТАТИСТИКА ===
        ws_stats = wb.create_sheet("Статистика")
        
        # Общая статистика
        stats_data = [
            ['Параметр', 'Значение'],
            ['Общее количество связей', len(df)],
            ['Дата создания', datetime.now().strftime("%Y-%m-%d %H:%M:%S")],
            ['', '']
        ]
        
        # Статистика по типам отношений
        if 'Relation_Type' in df.columns:
            stats_data.append(['Типы отношений', 'Количество'])
            relation_counts = df['Relation_Type'].value_counts()
            for relation_type, count in relation_counts.items():
                stats_data.append([relation_type, count])
            stats_data.append(['', ''])
        
        # Статистика схожести
        if 'Similarity_Score' in df.columns:
            stats_data.extend([
                ['Статистика схожести', ''],
                ['Средняя схожесть', f"{df['Similarity_Score'].mean():.3f}"],
                ['Максимальная схожесть', f"{df['Similarity_Score'].max():.3f}"],
                ['Минимальная схожесть', f"{df['Similarity_Score'].min():.3f}"],
                ['Медианная схожесть', f"{df['Similarity_Score'].median():.3f}"],
                ['', '']
            ])
        
        # Статистика по поисковым движкам
        if 'Search_Engine' in df.columns:
            stats_data.append(['Поисковые движки', 'Количество'])
            engine_counts = df['Search_Engine'].value_counts()
            for engine, count in engine_counts.items():
                stats_data.append([engine, count])
        
        # Записываем статистику
        for row_data in stats_data:
            ws_stats.append(row_data)
        
        # Форматирование статистики
        for col in range(1, 3):
            cell = ws_stats.cell(row=1, column=col)
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = header_alignment
        
        # Автоподбор ширины для статистики
        for column in ws_stats.columns:
            max_length = 0
            column_letter = column[0].column_letter
            
            for cell in column:
                try:
                    if len(str(cell.value)) > max_length:
                        max_length = len(str(cell.value))
                except:
                    pass
            
            adjusted_width = min(max_length + 2, 40)
            ws_stats.column_dimensions[column_letter].width = adjusted_width
        
        # === ЛИСТ 3: МЕТАДАННЫЕ ===
        if metadata:
            ws_meta = wb.create_sheet("Метаданные")
            
            meta_data = [['Параметр', 'Значение']]
            for key, value in metadata.items():
                meta_data.append([key, str(value)])
            
            for row_data in meta_data:
                ws_meta.append(row_data)
            
            # Форматирование метаданных
            for col in range(1, 3):
                cell = ws_meta.cell(row=1, column=col)
                cell.font = header_font
                cell.fill = header_fill
                cell.alignment = header_alignment
            
            # Автоподбор ширины для метаданных
            for column in ws_meta.columns:
                max_length = 0
                column_letter = column[0].column_letter
                
                for cell in column:
                    try:
                        if len(str(cell.value)) > max_length:
                            max_length = len(str(cell.value))
                    except:
                        pass
                
                adjusted_width = min(max_length + 2, 40)
                ws_meta.column_dimensions[column_letter].width = adjusted_width
        
        # Сохраняем файл
        wb.save(excel_file)
        logger.info(f"✅ Excel file created successfully: {excel_file}")
        
        # Показываем статистику
        logger.info(f"📊 File statistics:")
        logger.info(f"   Records: {len(df):,}")
        if 'Relation_Type' in df.columns:
            relation_counts = df['Relation_Type'].value_counts()
            for relation_type, count in relation_counts.items():
                logger.info(f"   {relation_type}: {count:,}")
        
        return True
        
    except Exception as e:
        logger.error(f"Error creating Excel file: {e}")
        return False


def main():
    parser = argparse.ArgumentParser(description='Convert CSV to Excel with formatting')
    parser.add_argument('csv_file', help='Input CSV file')
    parser.add_argument('-o', '--output', help='Output Excel file')
    parser.add_argument('--metadata', help='JSON metadata file')
    
    args = parser.parse_args()
    
    # Определяем выходной файл
    if args.output:
        excel_file = args.output
    else:
        csv_path = Path(args.csv_file)
        excel_file = csv_path.with_suffix('.xlsx')
    
    # Загружаем метаданные если есть
    metadata = None
    if args.metadata and os.path.exists(args.metadata):
        import json
        with open(args.metadata, 'r') as f:
            metadata = json.load(f)
    
    # Конвертируем
    success = create_excel_with_formatting(args.csv_file, excel_file, metadata)
    
    if success:
        print(f"✅ Conversion completed: {excel_file}")
        return 0
    else:
        print(f"❌ Conversion failed")
        return 1


if __name__ == "__main__":
    exit(main())
